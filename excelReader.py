from openpyxl import load_workbook
import  collections


#LOAD EXCEL AND SHEET
wb = load_workbook(filename = 'D:\pythonWorkspace\pythonXL\LinkTest.xlsx')
sheet_range = wb['Mainsheet']

#CREATE EMPTY LIST FOR CAPTURING KEYS AND VALUES
valueList = []
keyList = []

dict_x = collections.defaultdict(list)

#CREATE KEY VALUE PAIRS FROM EXCEL
for x in range(2, 5):
    b=sheet_range.cell(row=x,column=1).value
    for y in range(2, 4):
        a = sheet_range.cell(row=x, column=y).value
        #lst.append(a)
        dict_x[b].append(a)
        newobj = dict_x
        newobj.values()
print(newobj)


# READ VALUES FROM EACH KEY AS STRING
for key,val in dict_x.items():
    keyList = key
    print(keyList)
    valueList = (dict_x.get(keyList))
    for k in range(0, 2):
        cellval=valueList[k]
        print(valueList[k])










